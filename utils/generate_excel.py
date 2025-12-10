import os
import glob
import re
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

def parse_markdown_file(filepath):
    """
    Parses a markdown file, supporting multi-line indented sub-bullets for Action/Risk/Tools.
    """
    filename = os.path.basename(filepath).replace(".md", "")
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    items = []
    current_section = "General"
    current_item = None
    
    # State tracking: where are we currently appending text?
    # Options: None, 'Description', 'Risk', 'Action', 'Tools'
    current_mode = None 

    # Regex patterns
    section_pattern = re.compile(r'^##\s+(.+)')
    checklist_pattern = re.compile(r'^-\s+\[\s?\]\s+\*\*(.+?)\*\*[:?]?\s*(.*)')
    
    # Patterns to detect start of attributes
    # We look for lines starting with indent + - + *Key:*
    risk_start = re.compile(r'^\s+-\s+\*(Risk|Context|Risk/Context):\*\s*(.*)')
    action_start = re.compile(r'^\s+-\s+\*(Action):\*\s*(.*)')
    tools_start = re.compile(r'^\s+-\s+\*(Tools):\*\s*(.*)')
    
    # Pattern to detect sub-bullets (indent + - or * or number)
    sub_bullet_pattern = re.compile(r'^\s+[-*]\s+(.+)')

    for line in lines:
        line_content = line.strip()
        
        # Skip empty lines
        if not line_content:
            continue
        
        # 1. Detect Section Header
        section_match = section_pattern.match(line)
        if section_match:
            current_section = section_match.group(1).strip()
            current_mode = None
            continue

        # 2. Detect New Checklist Item
        item_match = checklist_pattern.match(line)
        if item_match:
            if current_item:
                items.append(current_item)
            
            current_item = {
                'Module': filename,
                'Section': current_section,
                'Check Item': item_match.group(1).strip(),
                'Description': item_match.group(2).strip(),
                'Risk/Context': "",
                'Action': "",
                'Tools': "",
                'Status': "" 
            }
            current_mode = 'Description' # Default mode after title
            continue

        # 3. Detect Attribute Starts (Risk, Action, Tools)
        if current_item:
            # Check Risk
            risk_match = risk_start.match(line)
            if risk_match:
                current_mode = 'Risk/Context'
                content = risk_match.group(2).strip()
                if content:
                    current_item['Risk/Context'] += content + "\n"
                continue

            # Check Action
            action_match = action_start.match(line)
            if action_match:
                current_mode = 'Action'
                content = action_match.group(2).strip()
                if content:
                    current_item['Action'] += content + "\n"
                continue
                
            # Check Tools
            tools_match = tools_start.match(line)
            if tools_match:
                current_mode = 'Tools'
                content = tools_match.group(2).strip()
                if content:
                    current_item['Tools'] += content + "\n"
                continue

            # 4. Handle Sub-bullets / Continuation Lines
            # If the line is indented and we are inside a mode (Action/Risk/Tools)
            if current_mode in ['Risk/Context', 'Action', 'Tools']:
                # Check if it's a bullet point
                sub_match = sub_bullet_pattern.match(line)
                if sub_match:
                    # Add a clean bullet point for Excel
                    current_item[current_mode] += "• " + sub_match.group(1).strip() + "\n"
                else:
                    # It might be just a wrapped line of text, append it
                    # But ensure it's indented in the original file (to avoid capturing next items)
                    if line.startswith(" ") or line.startswith("\t"):
                         current_item[current_mode] += line.strip() + " "

    # Append the last item found
    if current_item:
        items.append(current_item)

    return items

def generate_excel(input_dir="checklists", output_file="GeoAI_Checklist.xlsx"):
    print(f"🔍 Scanning {input_dir} for Markdown files...")
    md_files = sorted(glob.glob(os.path.join(input_dir, "*.md")))
    
    all_data = []
    for md_file in md_files:
        print(f"   Parsing {os.path.basename(md_file)}...")
        all_data.extend(parse_markdown_file(md_file))

    if not all_data:
        print("❌ No checklist items found. Please check markdown formatting.")
        return

    df = pd.DataFrame(all_data)
    
    # Cleanup: Strip trailing newlines
    cols_to_clean = ['Risk/Context', 'Action', 'Tools']
    for col in cols_to_clean:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    # Reorder
    final_cols = ['Module', 'Section', 'Check Item', 'Description', 'Risk/Context', 'Action', 'Tools', 'Status']
    # Ensure columns exist even if empty
    for col in final_cols:
        if col not in df.columns:
            df[col] = ""
            
    df = df[final_cols]

    print(f"💾 Saving to {output_file}...")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Checklist')
        
        workbook = writer.book
        worksheet = writer.sheets['Checklist']
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid") # Dark Blue
        alignment = Alignment(wrap_text=True, vertical='top')
        
        # Set Column Widths
        worksheet.column_dimensions['A'].width = 15 # Module
        worksheet.column_dimensions['B'].width = 20 # Section
        worksheet.column_dimensions['C'].width = 25 # Check Item
        worksheet.column_dimensions['D'].width = 30 # Description
        worksheet.column_dimensions['E'].width = 45 # Risk
        worksheet.column_dimensions['F'].width = 50 # Action (Wider for bullet points)
        worksheet.column_dimensions['G'].width = 35 # Tools
        worksheet.column_dimensions['H'].width = 12 # Status

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = alignment
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill

    print(f"✅ Done! Created {output_file} with {len(df)} items.")

if __name__ == "__main__":
    generate_excel()